import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Personal Acrópolis"

# Estilos
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="0F1729", end_color="0F1729", fill_type="solid")
accent_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
accent_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
data_font = Font(name="Calibri", size=11)
border = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")

# Título
ws.merge_cells("A1:F1")
title_cell = ws["A1"]
title_cell.value = "PLANILLA DE PERSONAL — COLEGIO ACRÓPOLIS"
title_cell.font = accent_font
title_cell.fill = accent_fill
title_cell.alignment = center

# Subtítulo
ws.merge_cells("A2:F2")
sub = ws["A2"]
sub.value = "Importar en: Personal → Importar Excel"
sub.font = Font(name="Calibri", italic=True, size=10, color="6B7280")
sub.alignment = center

# Headers (fila 4)
headers = ["RUN", "APELLIDOS", "NOMBRE", "CARGO", "EMAIL", "MINUTOS CONTRATO"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

# Datos de ejemplo (20 funcionarios)
datos = [
    ("12.345.678-9", "ALLENDE ROJAS",     "MARCELA",    "Docente",         "m.allende@acropolis.cl",    0),
    ("13.456.789-0", "ARAYA MUÑOZ",       "CAROLINA",   "Docente",         "c.araya@acropolis.cl",      0),
    ("14.567.890-1", "CONTRERAS SILVA",    "PAMELA",     "Docente",         "p.contreras@acropolis.cl",  0),
    ("15.678.901-2", "DÍAZ RAMÍREZ",       "ALEJANDRA",  "Docente",         "a.diaz@acropolis.cl",       45),
    ("16.789.012-3", "ESPINOZA CASTILLO",  "MARÍA",      "Docente",         "m.espinoza@acropolis.cl",   0),
    ("17.890.123-4", "FUENTES LÓPEZ",      "JORGE",      "Docente",         "j.fuentes@acropolis.cl",    90),
    ("18.901.234-5", "GÓMEZ HERRERA",      "CLAUDIA",    "Asist. Educación","c.gomez@acropolis.cl",      0),
    ("19.012.345-6", "HERNÁNDEZ VEGA",     "ROBERTO",    "Docente",         "r.hernandez@acropolis.cl",  0),
    ("20.123.456-7", "IBARRA TORRES",      "FRANCISCA",  "Docente",         "f.ibarra@acropolis.cl",     45),
    ("21.234.567-8", "JARA MORALES",       "CRISTIAN",   "Docente",         "c.jara@acropolis.cl",       0),
    ("22.345.678-9", "LAGOS PIZARRO",      "VERÓNICA",   "UTP",             "v.lagos@acropolis.cl",      0),
    ("23.456.789-0", "MELLA SEPÚLVEDA",    "PATRICIA",   "Docente",         "p.mella@acropolis.cl",      90),
    ("24.567.890-1", "NAVARRO BRAVO",      "GONZALO",    "Docente",         "g.navarro@acropolis.cl",    0),
    ("25.678.901-2", "ORTEGA REYES",       "DANIELA",    "Docente",         "d.ortega@acropolis.cl",     0),
    ("26.789.012-3", "PAREDES SOTO",       "LUIS",       "Docente",         "l.paredes@acropolis.cl",    45),
    ("27.890.123-4", "QUINTANA FIGUEROA",  "ANDREA",     "Docente",         "a.quintana@acropolis.cl",   0),
    ("28.901.234-5", "RÍOS VALENZUELA",    "FERNANDO",   "Inspector",       "f.rios@acropolis.cl",       0),
    ("29.012.345-6", "SALAZAR GUZMÁN",     "NATALIA",    "Docente",         "n.salazar@acropolis.cl",    0),
    ("30.123.456-7", "TAPIA OLIVARES",     "MIGUEL",     "Docente",         "m.tapia@acropolis.cl",      90),
    ("31.234.567-8", "URZÚA MENDOZA",      "CAMILA",     "Orientadora",     "c.urzua@acropolis.cl",      0),
]

even_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

for row_idx, fila in enumerate(datos, 5):
    for col_idx, valor in enumerate(fila, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=valor)
        cell.font = data_font
        cell.border = border
        cell.alignment = center if col_idx in (1, 6) else left
        if row_idx % 2 == 0:
            cell.fill = even_fill

# Anchos
ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 20
ws.column_dimensions["E"].width = 28
ws.column_dimensions["F"].width = 20

# Nota
note_row = len(datos) + 6
ws.merge_cells(f"A{note_row}:F{note_row}")
note = ws.cell(row=note_row, column=1)
note.value = "Puede agregar mas filas. Las columnas se detectan automaticamente por encabezado."
note.font = Font(name="Calibri", italic=True, size=9, color="9CA3AF")
note.alignment = Alignment(horizontal="center")

path = r"E:\ANTIGRAVITY\REEMPLAZOS Y AUSENCIAS\plantilla_personal_acropolis.xlsx"
wb.save(path)
print(f"Plantilla generada: {path}")
